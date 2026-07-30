from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg
from .models import (
    Car, ContactMessage, Review, Brand, Client,
    RentalContract, TechnicalInspection, Insurance
)
import csv
import json
import datetime


# ============================================
#  PAGES PUBLIQUES
# ============================================

def home(request):
    """Page d'accueil avec hero section et aperçu des véhicules."""
    featured_cars = Car.objects.filter(is_available=True, featured=True)[:6]

    approved_reviews = Review.objects.filter(is_approved=True)[:6]

    context = {
        'featured_cars': featured_cars,
        'reviews': approved_reviews,
        'page_title': 'SAOUD CAR — Location de Voitures à Ben Slimane',
    }
    return render(request, 'core/home.html', context)


def cars(request):
    """Page catalogue des véhicules avec filtrage."""
    category = request.GET.get('category', '')
    if category:
        vehicles = Car.objects.filter(is_available=True, category=category)
    else:
        vehicles = Car.objects.filter(is_available=True)

    context = {
        'vehicles': vehicles,
        'categories': Car.CATEGORY_CHOICES,
        'selected_category': category,
        'page_title': 'Nos Véhicules — SAOUD CAR',
    }
    return render(request, 'core/cars.html', context)


def about(request):
    """Page à propos de l'entreprise."""
    context = {
        'page_title': 'À Propos — SAOUD CAR',
    }
    return render(request, 'core/about.html', context)


def contact(request):
    """Page de contact avec formulaire."""
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        subject = request.POST.get('subject', '')
        message_text = request.POST.get('message', '')

        if name and email and subject and message_text:
            ContactMessage.objects.create(
                name=name,
                email=email,
                phone=phone,
                subject=subject,
                message=message_text,
            )
            
            # Créer automatiquement un prospect dans la table Client s'il n'existe pas
            from django.db.models import Q
            from .models import Client
            import time
            
            # Chercher si un client avec cet email ou ce téléphone existe déjà
            existing_client = None
            if email or phone:
                query = Q()
                if email:
                    query |= Q(email=email)
                if phone:
                    query |= Q(phone=phone)
                existing_client = Client.objects.filter(query).first()
                
            if not existing_client:
                # Générer un CIN temporaire pour le prospect
                placeholder_cin = f"PROSPECT-{int(time.time())}"
                Client.objects.create(
                    full_name=name,
                    email=email,
                    phone=phone,
                    cin=placeholder_cin,
                    address="Via Formulaire de Contact"
                )
            messages.success(request, 'Votre message a été envoyé avec succès ! Nous vous répondrons dans les plus brefs délais.')
            return redirect('core:contact')
        else:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')

    context = {
        'page_title': 'Contact — SAOUD CAR',
    }
    return render(request, 'core/contact.html', context)


def submit_review(request):
    """Soumettre un avis client."""
    if request.method == 'POST':
        name = request.POST.get('name', '')
        city = request.POST.get('city', '')
        rating = int(request.POST.get('rating', 5))
        comment = request.POST.get('comment', '')

        if name and comment and 1 <= rating <= 5:
            Review.objects.create(
                name=name,
                city=city,
                rating=rating,
                comment=comment,
                is_approved=False,
            )
            messages.success(request, 'Merci pour votre avis ! Il sera publié après validation.')
            return redirect('core:home')
        else:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')

    return redirect('core:home')


# ============================================
#  DASHBOARD ADMIN PERSONNALISÉ
# ============================================

def admin_dashboard(request):
    """Page d'accueil du dashboard admin avec graphiques."""
    today = datetime.date.today()
    current_month = today.month
    current_year = today.year

    # Stats de base
    total_cars = Car.objects.count()
    available_cars = Car.objects.filter(is_available=True).count()
    total_clients = Client.objects.count()
    active_contracts = RentalContract.objects.filter(status='en_cours').count()
    total_reviews = Review.objects.count()
    pending_reviews = Review.objects.filter(is_approved=False).count()
    total_messages = ContactMessage.objects.count()
    unread_messages = ContactMessage.objects.filter(is_read=False).count()

    # Revenus du mois en cours
    month_revenue = RentalContract.objects.filter(
        status='termine',
        end_date__year=current_year,
        end_date__month=current_month
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Revenus de l'année
    year_revenue = RentalContract.objects.filter(
        status='termine',
        end_date__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Données pour le graphique de revenus mensuels (12 derniers mois)
    monthly_revenue_data = []
    monthly_labels = []
    for i in range(11, -1, -1):
        d = today - datetime.timedelta(days=i * 30)
        month = d.month
        year = d.year
        rev = RentalContract.objects.filter(
            status='termine',
            end_date__year=year,
            end_date__month=month
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        monthly_revenue_data.append(float(rev))
        month_names = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
        monthly_labels.append(f"{month_names[month - 1]} {year}")

    # Locations par mois (Courbe)
    monthly_rentals_data = []
    for i in range(11, -1, -1):
        d = today - datetime.timedelta(days=i * 30)
        month = d.month
        year = d.year
        rentals_count = RentalContract.objects.filter(
            start_date__year=year,
            start_date__month=month
        ).count()
        monthly_rentals_data.append(rentals_count)

    # Revenus par Marque (Bâtons)
    brand_revenue_labels = []
    brand_revenue_data = []
    # On récupère toutes les marques (ou les noms de marques via car__brand_ref__name)
    brands_data = RentalContract.objects.filter(status='termine').values('car__brand_ref__name').annotate(total=Sum('total_amount')).order_by('-total')[:5]
    for b in brands_data:
        name = b['car__brand_ref__name'] or 'Inconnue'
        total = b['total'] or 0
        brand_revenue_labels.append(name)
        brand_revenue_data.append(float(total))

    # Statut de la flotte (Camembert)
    fleet_status_labels = ['Disponibles', 'En Location']
    rented_cars = Car.objects.filter(is_available=False).count()
    # or actually we can just use available_cars and (total_cars - available_cars)
    fleet_status_data = [available_cars, total_cars - available_cars]

    # Données pour le graphique par catégorie
    category_data = []
    category_labels = []
    for code, label in Car.CATEGORY_CHOICES:
        count = RentalContract.objects.filter(car__category=code).count()
        if count > 0:
            category_data.append(count)
            category_labels.append(label)

    # Top 3 véhicules les plus loués
    top_cars = Car.objects.annotate(
        rental_count=Count('contracts')
    ).order_by('-rental_count')[:3]

    # Alertes (visites et assurances qui expirent bientôt)
    alert_date = today + datetime.timedelta(days=30)
    expiring_inspections = TechnicalInspection.objects.filter(
        expiry_date__lte=alert_date
    ).select_related('car')
    expiring_insurances = Insurance.objects.filter(
        expiry_date__lte=alert_date
    ).select_related('car')

    # ========= NOUVEAUX GRAPHIQUES =========

    # Répartition par type de carburant (Camembert)
    fuel_labels = []
    fuel_data = []
    for code, label in [('essence', 'Essence'), ('diesel', 'Diesel'), ('hybride', 'Hybride')]:
        count = Car.objects.filter(fuel_type=code).count()
        if count > 0:
            fuel_labels.append(label)
            fuel_data.append(count)

    # Répartition par transmission (Camembert)
    trans_labels = []
    trans_data = []
    for code, label in [('manuelle', 'Manuelle'), ('automatique', 'Automatique')]:
        count = Car.objects.filter(transmission=code).count()
        if count > 0:
            trans_labels.append(label)
            trans_data.append(count)

    # Voitures par catégorie (Bâtons)
    cars_by_category_labels = []
    cars_by_category_data = []
    for code, label in Car.CATEGORY_CHOICES:
        count = Car.objects.filter(category=code).count()
        cars_by_category_labels.append(label)
        cars_by_category_data.append(count)

    # Revenus + Locations combinés (Double axe)
    # On réutilise monthly_revenue_data et monthly_rentals_data

    # Total contrats terminés
    total_completed = RentalContract.objects.filter(status='termine').count()
    total_contracts = RentalContract.objects.count()

    # Taux d'occupation de la flotte
    occupation_rate = round((total_cars - available_cars) / total_cars * 100, 1) if total_cars > 0 else 0

    # Revenu moyen par contrat
    avg_revenue = 0
    if total_completed > 0:
        avg_rev_qs = RentalContract.objects.filter(status='termine').aggregate(avg=Avg('total_amount'))
        avg_revenue = float(avg_rev_qs['avg'] or 0)

    context = {
        'total_cars': total_cars,
        'available_cars': available_cars,
        'total_clients': total_clients,
        'active_contracts': active_contracts,
        'total_reviews': total_reviews,
        'pending_reviews': pending_reviews,
        'total_messages': total_messages,
        'unread_messages': unread_messages,
        'month_revenue': month_revenue,
        'year_revenue': year_revenue,
        'monthly_revenue_data': json.dumps(monthly_revenue_data),
        'monthly_rentals_data': json.dumps(monthly_rentals_data),
        'monthly_labels': json.dumps(monthly_labels),
        'category_data': json.dumps(category_data),
        'category_labels': json.dumps(category_labels),
        'brand_revenue_labels': json.dumps(brand_revenue_labels),
        'brand_revenue_data': json.dumps(brand_revenue_data),
        'fleet_status_labels': json.dumps(fleet_status_labels),
        'fleet_status_data': json.dumps(fleet_status_data),
        # Nouvelles données
        'fuel_labels': json.dumps(fuel_labels),
        'fuel_data': json.dumps(fuel_data),
        'trans_labels': json.dumps(trans_labels),
        'trans_data': json.dumps(trans_data),
        'cars_by_category_labels': json.dumps(cars_by_category_labels),
        'cars_by_category_data': json.dumps(cars_by_category_data),
        'total_completed': total_completed,
        'total_contracts': total_contracts,
        'occupation_rate': occupation_rate,
        'avg_revenue': avg_revenue,
        'rented_cars_list': Car.objects.filter(is_available=False).order_by('expected_return_date'),
        'top_cars': top_cars,
        'expiring_inspections': expiring_inspections,
        'expiring_insurances': expiring_insurances,
        'recent_reviews': Review.objects.all()[:5],
        'recent_messages': ContactMessage.objects.all()[:5],
    }
    return render(request, 'core/admin/dashboard.html', context)


# ============================================
#  ADMIN — VOITURES
# ============================================

def admin_cars(request):
    """Liste des voitures avec gestion rapide du statut."""
    if request.method == 'POST':
        action = request.POST.get('action')
        car_id = request.POST.get('car_id')
        if action and car_id:
            car = get_object_or_404(Car, id=car_id)
            if action == 'mark_rented':
                return_date = request.POST.get('return_date')
                if return_date:
                    car.is_available = False
                    car.expected_return_date = return_date
                    car.save()
                    messages.success(request, f'{car.brand.name} {car.model} marquée en location.')
                else:
                    messages.error(request, 'Veuillez préciser une date de retour prévue.')
            elif action == 'mark_available':
                car.is_available = True
                car.expected_return_date = None
                car.save()
                messages.success(request, f'{car.brand.name} {car.model} marquée comme disponible.')
        return redirect('core:admin_cars')

    cars = Car.objects.all().order_by('-id')
    context = {'cars': cars}
    return render(request, 'core/admin/cars_list.html', context)


def admin_car_add(request):
    """Ajouter une voiture."""
    if request.method == 'POST':
        car = Car(
            name=request.POST.get('name', ''),
            brand=request.POST.get('brand', ''),
            category=request.POST.get('category', 'economique'),
            price_per_day=request.POST.get('price_per_day', 0),
            seats=request.POST.get('seats', 5),
            transmission=request.POST.get('transmission', 'manuelle'),
            fuel_type=request.POST.get('fuel_type', 'diesel'),
            license_plate=request.POST.get('license_plate', ''),
            year=request.POST.get('year') or None,
            mileage=request.POST.get('mileage', 0) or 0,
            color=request.POST.get('color', ''),
            is_available='is_available' in request.POST,
            featured='featured' in request.POST,
            description=request.POST.get('description', ''),
        )
        # Link to brand FK if brand_ref is selected
        brand_ref_id = request.POST.get('brand_ref')
        if brand_ref_id:
            car.brand_ref_id = int(brand_ref_id)
            car.brand = Brand.objects.get(id=int(brand_ref_id)).name
        if request.FILES.get('image'):
            car.image = request.FILES['image']
        car.save()
        messages.success(request, f'Véhicule "{car.get_brand_display()} {car.name}" ajouté avec succès !')
        return redirect('core:admin_cars')

    context = {
        'categories': Car.CATEGORY_CHOICES,
        'transmissions': [('manuelle', 'Manuelle'), ('automatique', 'Automatique')],
        'fuel_types': [('essence', 'Essence'), ('diesel', 'Diesel'), ('hybride', 'Hybride')],
        'brands': Brand.objects.all(),
    }
    return render(request, 'core/admin/car_form.html', context)


def admin_car_edit(request, car_id):
    """Modifier une voiture."""
    car = get_object_or_404(Car, id=car_id)

    if request.method == 'POST':
        car.name = request.POST.get('name', car.name)
        car.brand = request.POST.get('brand', car.brand)
        car.category = request.POST.get('category', car.category)
        car.price_per_day = request.POST.get('price_per_day', car.price_per_day)
        car.seats = request.POST.get('seats', car.seats)
        car.transmission = request.POST.get('transmission', car.transmission)
        car.fuel_type = request.POST.get('fuel_type', car.fuel_type)
        car.license_plate = request.POST.get('license_plate', '')
        car.year = request.POST.get('year') or None
        car.mileage = request.POST.get('mileage', 0) or 0
        car.color = request.POST.get('color', '')
        car.is_available = 'is_available' in request.POST
        car.featured = 'featured' in request.POST
        car.description = request.POST.get('description', '')
        brand_ref_id = request.POST.get('brand_ref')
        if brand_ref_id:
            car.brand_ref_id = int(brand_ref_id)
            car.brand = Brand.objects.get(id=int(brand_ref_id)).name
        else:
            car.brand_ref = None
        if request.FILES.get('image'):
            car.image = request.FILES['image']
        car.save()
        messages.success(request, f'Véhicule "{car.get_brand_display()} {car.name}" modifié avec succès !')
        return redirect('core:admin_cars')

    context = {
        'car': car,
        'categories': Car.CATEGORY_CHOICES,
        'transmissions': [('manuelle', 'Manuelle'), ('automatique', 'Automatique')],
        'fuel_types': [('essence', 'Essence'), ('diesel', 'Diesel'), ('hybride', 'Hybride')],
        'brands': Brand.objects.all(),
    }
    return render(request, 'core/admin/car_form.html', context)


def admin_car_delete(request, car_id):
    """Supprimer une voiture."""
    car = get_object_or_404(Car, id=car_id)
    name = f"{car.get_brand_display()} {car.name}"
    car.delete()
    messages.success(request, f'Véhicule "{name}" supprimé avec succès !')
    return redirect('core:admin_cars')


def admin_car_toggle_featured(request, car_id):
    """Afficher ou masquer une voiture de la page d'accueil."""
    car = get_object_or_404(Car, id=car_id)
    car.featured = not car.featured
    car.save()
    status = "affiché en page d'accueil" if car.featured else "retiré de la page d'accueil"
    messages.success(request, f'"{car.get_brand_display()} {car.name}" {status}.')
    return redirect('core:admin_cars')


# ============================================
#  ADMIN — MARQUES
# ============================================

def admin_brands(request):
    """Liste des marques."""
    all_brands = Brand.objects.annotate(
        total_cars=Count('cars'),
        available_count=Count('cars', filter=Q(cars__is_available=True)),
        rented_count=Count('cars', filter=Q(cars__is_available=False))
    ).all()
    context = {'all_brands': all_brands}
    return render(request, 'core/admin/brands_list.html', context)


def admin_brand_add(request):
    """Ajouter une marque."""
    if request.method == 'POST':
        brand = Brand(
            name=request.POST.get('name', ''),
            country=request.POST.get('country', ''),
        )
        if request.FILES.get('logo'):
            brand.logo = request.FILES['logo']
        brand.save()
        messages.success(request, f'Marque "{brand.name}" ajoutée avec succès !')
        return redirect('core:admin_brands')
    return render(request, 'core/admin/brand_form.html')


def admin_brand_edit(request, brand_id):
    """Modifier une marque."""
    brand = get_object_or_404(Brand, id=brand_id)
    if request.method == 'POST':
        brand.name = request.POST.get('name', brand.name)
        brand.country = request.POST.get('country', brand.country)
        if request.FILES.get('logo'):
            brand.logo = request.FILES['logo']
            brand.logo_url = ''  # Clear external URL if a new file is uploaded
        brand.save()
        messages.success(request, f'Marque "{brand.name}" modifiée avec succès !')
        return redirect('core:admin_brands')
    context = {'brand': brand}
    return render(request, 'core/admin/brand_form.html', context)


def admin_brand_delete(request, brand_id):
    """Supprimer une marque."""
    brand = get_object_or_404(Brand, id=brand_id)
    name = brand.name
    brand.delete()
    messages.success(request, f'Marque "{name}" supprimée avec succès !')
    return redirect('core:admin_brands')


# ============================================
#  ADMIN — CLIENTS
# ============================================

def admin_clients(request):
    """Liste des clients."""
    all_clients = Client.objects.all()
    context = {'all_clients': all_clients}
    return render(request, 'core/admin/clients_list.html', context)


def admin_client_add(request):
    """Ajouter un client."""
    if request.method == 'POST':
        client = Client(
            full_name=request.POST.get('full_name', ''),
            cin=request.POST.get('cin', ''),
            drivers_license=request.POST.get('drivers_license', ''),
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            address=request.POST.get('address', ''),
            city=request.POST.get('city', ''),
        )
        if request.FILES.get('cin_front'):
            client.cin_front = request.FILES['cin_front']
        if request.FILES.get('cin_back'):
            client.cin_back = request.FILES['cin_back']
        client.save()
        messages.success(request, f'Client "{client.full_name}" ajouté avec succès !')
        return redirect('core:admin_clients')
    return render(request, 'core/admin/client_form.html')


def admin_client_edit(request, client_id):
    """Modifier un client."""
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        client.full_name = request.POST.get('full_name', client.full_name)
        client.cin = request.POST.get('cin', client.cin)
        client.drivers_license = request.POST.get('drivers_license', client.drivers_license)
        client.phone = request.POST.get('phone', client.phone)
        client.email = request.POST.get('email', client.email)
        client.address = request.POST.get('address', client.address)
        client.city = request.POST.get('city', client.city)
        if request.FILES.get('cin_front'):
            client.cin_front = request.FILES['cin_front']
        if request.FILES.get('cin_back'):
            client.cin_back = request.FILES['cin_back']
        client.save()
        messages.success(request, f'Client "{client.full_name}" modifié avec succès !')
        return redirect('core:admin_clients')
    context = {'client': client}
    return render(request, 'core/admin/client_form.html', context)


def admin_client_delete(request, client_id):
    """Supprimer un client."""
    client = get_object_or_404(Client, id=client_id)
    name = client.full_name
    client.delete()
    messages.success(request, f'Client "{name}" supprimé avec succès !')
    return redirect('core:admin_clients')


# ============================================
#  ADMIN — CONTRATS
# ============================================

def admin_contracts(request):
    """Liste des contrats avec filtres."""
    contracts = RentalContract.objects.select_related('client', 'car').all()

    # Filtres
    status_filter = request.GET.get('status', '')
    if status_filter:
        contracts = contracts.filter(status=status_filter)

    context = {
        'all_contracts': contracts,
        'status_filter': status_filter,
    }
    return render(request, 'core/admin/contracts_list.html', context)


def admin_contract_add(request):
    """Ajouter un contrat."""
    if request.method == 'POST':
        contract = RentalContract(
            client_id=int(request.POST.get('client', 0)),
            car_id=int(request.POST.get('car', 0)),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date'),
            price_per_day=request.POST.get('price_per_day', 0),
            deposit=request.POST.get('deposit', 0) or 0,
            km_start=request.POST.get('km_start', 0) or 0,
            km_end=request.POST.get('km_end') or None,
            status=request.POST.get('status', 'en_cours'),
            notes=request.POST.get('notes', ''),
        )
        contract.save()
        messages.success(request, f'Contrat {contract.contract_number} créé avec succès !')
        return redirect('core:admin_contracts')

    context = {
        'clients': Client.objects.all(),
        'cars': Car.objects.all(),
        'statuses': RentalContract.STATUS_CHOICES,
    }
    return render(request, 'core/admin/contract_form.html', context)


def admin_contract_edit(request, contract_id):
    """Modifier un contrat."""
    contract = get_object_or_404(RentalContract, id=contract_id)
    if request.method == 'POST':
        contract.client_id = int(request.POST.get('client', contract.client_id))
        contract.car_id = int(request.POST.get('car', contract.car_id))
        contract.start_date = request.POST.get('start_date', contract.start_date)
        contract.end_date = request.POST.get('end_date', contract.end_date)
        contract.price_per_day = request.POST.get('price_per_day', contract.price_per_day)
        contract.deposit = request.POST.get('deposit', contract.deposit) or 0
        contract.km_start = request.POST.get('km_start', contract.km_start) or 0
        contract.km_end = request.POST.get('km_end') or None
        contract.status = request.POST.get('status', contract.status)
        contract.notes = request.POST.get('notes', contract.notes)
        contract.save()
        messages.success(request, f'Contrat {contract.contract_number} modifié avec succès !')
        return redirect('core:admin_contracts')

    context = {
        'contract': contract,
        'clients': Client.objects.all(),
        'cars': Car.objects.all(),
        'statuses': RentalContract.STATUS_CHOICES,
    }
    return render(request, 'core/admin/contract_form.html', context)


def admin_contract_delete(request, contract_id):
    """Supprimer un contrat."""
    contract = get_object_or_404(RentalContract, id=contract_id)
    num = contract.contract_number
    contract.delete()
    messages.success(request, f'Contrat {num} supprimé avec succès !')
    return redirect('core:admin_contracts')


def admin_contract_pdf(request, contract_id):
    """Générer un PDF pour un contrat."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io

    contract = get_object_or_404(
        RentalContract.objects.select_related('client', 'car'),
        id=contract_id
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#E53935'), spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=20)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1A1A2E'), spaceBefore=16, spaceAfter=8)
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=10, leading=14)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=colors.grey, leading=12)

    # Header
    elements.append(Paragraph("SAOUD CAR", title_style))
    elements.append(Paragraph("Location de Voitures — Ben Slimane, Maroc", subtitle_style))
    elements.append(Paragraph(f"CONTRAT DE LOCATION N° {contract.contract_number}", ParagraphStyle('ContractNum', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER, textColor=colors.HexColor('#1A1A2E'), spaceAfter=20)))

    # Client info
    elements.append(Paragraph("INFORMATIONS DU CLIENT", heading_style))
    client_data = [
        ['Nom complet', contract.client.full_name, 'CIN', contract.client.cin],
        ['Permis', contract.client.drivers_license or '—', 'Téléphone', contract.client.phone],
        ['Email', contract.client.email or '—', 'Ville', contract.client.city or '—'],
    ]
    client_table = Table(client_data, colWidths=[3.5 * cm, 5.5 * cm, 3.5 * cm, 5.5 * cm])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F5F5F5')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(client_table)

    # Vehicle info
    elements.append(Paragraph("INFORMATIONS DU VÉHICULE", heading_style))
    car = contract.car
    vehicle_data = [
        ['Véhicule', f"{car.get_brand_display()} {car.name}", 'Catégorie', car.get_category_display()],
        ['Immatriculation', car.license_plate or '—', 'Carburant', car.get_fuel_type_display()],
        ['Transmission', car.get_transmission_display(), 'Couleur', car.color or '—'],
    ]
    vehicle_table = Table(vehicle_data, colWidths=[3.5 * cm, 5.5 * cm, 3.5 * cm, 5.5 * cm])
    vehicle_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F5F5F5')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(vehicle_table)

    # Rental details
    elements.append(Paragraph("DÉTAILS DE LA LOCATION", heading_style))
    rental_data = [
        ['Date de début', str(contract.start_date.strftime('%d/%m/%Y')), 'Date de fin', str(contract.end_date.strftime('%d/%m/%Y'))],
        ['Durée', f"{contract.duration_days} jour(s)", 'Km départ', str(contract.km_start)],
        ['Km retour', str(contract.km_end or '—'), 'Statut', contract.get_status_display()],
    ]
    rental_table = Table(rental_data, colWidths=[3.5 * cm, 5.5 * cm, 3.5 * cm, 5.5 * cm])
    rental_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F5F5F5')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(rental_table)

    # Pricing
    elements.append(Paragraph("TARIFICATION", heading_style))
    price_data = [
        ['Prix par jour', f"{contract.price_per_day} MAD"],
        ['Durée', f"{contract.duration_days} jour(s)"],
        ['Caution', f"{contract.deposit} MAD"],
        ['MONTANT TOTAL', f"{contract.total_amount} MAD"],
    ]
    price_table = Table(price_data, colWidths=[9 * cm, 9 * cm])
    price_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('BACKGROUND', (-2, -1), (-1, -1), colors.HexColor('#E53935')),
        ('TEXTCOLOR', (-2, -1), (-1, -1), colors.white),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (-2, -1), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('RIGHTPADDING', (1, 0), (1, -1), 10),
    ]))
    elements.append(price_table)

    # Terms
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("CONDITIONS GÉNÉRALES", heading_style))
    terms = [
        "Le locataire s'engage à restituer le véhicule dans l'état dans lequel il l'a reçu.",
        "Tout dommage constaté au retour sera à la charge du locataire.",
        "La caution sera restituée après vérification de l'état du véhicule.",
        "Le carburant n'est pas inclus dans le prix de la location.",
        "Le locataire doit respecter le code de la route marocain.",
    ]
    for i, term in enumerate(terms, 1):
        elements.append(Paragraph(f"{i}. {term}", normal_style))

    # Signatures
    elements.append(Spacer(1, 40))
    sig_data = [
        ['Signature du Loueur', '', 'Signature du Locataire'],
        ['SAOUD CAR', '', contract.client.full_name],
        ['', '', ''],
        ['', '', ''],
        ['___________________', '', '___________________'],
    ]
    sig_table = Table(sig_data, colWidths=[7 * cm, 4 * cm, 7 * cm])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sig_table)

    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"Contrat généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} — SAOUD CAR, Hay Lalla Meriem, Ben Slimane — Tél: +212 661 395 495",
        small_style
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="contrat_{contract.contract_number}.pdf"'
    return response


def admin_contracts_export_csv(request):
    """Export des contrats en CSV."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="contrats_saoudcar.csv"'
    response.write('\ufeff')  # UTF-8 BOM for Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'N° Contrat', 'Client', 'CIN', 'Téléphone', 'Voiture',
        'Date début', 'Date fin', 'Durée (jours)', 'Prix/jour (MAD)',
        'Total (MAD)', 'Caution (MAD)', 'Statut'
    ])

    contracts = RentalContract.objects.select_related('client', 'car').all()
    for c in contracts:
        writer.writerow([
            c.contract_number,
            c.client.full_name,
            c.client.cin,
            c.client.phone,
            f"{c.car.get_brand_display()} {c.car.name}",
            c.start_date.strftime('%d/%m/%Y'),
            c.end_date.strftime('%d/%m/%Y'),
            c.duration_days,
            c.price_per_day,
            c.total_amount,
            c.deposit,
            c.get_status_display(),
        ])

    return response


def admin_contracts_export_excel(request):
    """Export des contrats en Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Contrats SAOUD CAR"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'N° Contrat', 'Client', 'CIN', 'Téléphone', 'Voiture',
        'Date début', 'Date fin', 'Durée (jours)', 'Prix/jour (MAD)',
        'Total (MAD)', 'Caution (MAD)', 'Statut'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    contracts = RentalContract.objects.select_related('client', 'car').all()
    for row_idx, c in enumerate(contracts, 2):
        data = [
            c.contract_number,
            c.client.full_name,
            c.client.cin,
            c.client.phone,
            f"{c.car.get_brand_display()} {c.car.name}",
            c.start_date.strftime('%d/%m/%Y'),
            c.end_date.strftime('%d/%m/%Y'),
            c.duration_days,
            float(c.price_per_day),
            float(c.total_amount),
            float(c.deposit),
            c.get_status_display(),
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 30)

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contrats_saoudcar.xlsx"'
    return response


# ============================================
#  ADMIN — VISITES TECHNIQUES
# ============================================

def admin_inspections(request):
    """Liste des visites techniques par véhicule."""
    cars = Car.objects.all().prefetch_related('inspections')
    car_list = []
    today = datetime.date.today()
    for car in cars:
        latest = car.inspections.order_by('-expiry_date').first()
        is_valid = latest.expiry_date >= today if latest and latest.expiry_date else False
        car_list.append({
            'car': car,
            'latest_inspection': latest,
            'is_valid': is_valid
        })
    context = {'car_list': car_list}
    return render(request, 'core/admin/inspections_list.html', context)


def admin_inspection_add(request):
    """Ajouter une visite technique."""
    if request.method == 'POST':
        inspection = TechnicalInspection(
            car_id=int(request.POST.get('car', 0)),
            inspection_date=request.POST.get('inspection_date'),
            expiry_date=request.POST.get('expiry_date'),
            result=request.POST.get('result', 'favorable'),
            center_name=request.POST.get('center_name', ''),
            cost=request.POST.get('cost', 0) or 0,
        )
        if request.FILES.get('document'):
            inspection.document = request.FILES['document']
        inspection.save()
        messages.success(request, 'Visite technique ajoutée avec succès !')
        return redirect('core:admin_inspections')

    context = {'cars': Car.objects.all()}
    return render(request, 'core/admin/inspection_form.html', context)


def admin_inspection_edit(request, inspection_id):
    """Modifier une visite technique."""
    inspection = get_object_or_404(TechnicalInspection, id=inspection_id)
    if request.method == 'POST':
        inspection.car_id = int(request.POST.get('car', inspection.car_id))
        inspection.inspection_date = request.POST.get('inspection_date', inspection.inspection_date)
        inspection.expiry_date = request.POST.get('expiry_date', inspection.expiry_date)
        inspection.result = request.POST.get('result', inspection.result)
        inspection.center_name = request.POST.get('center_name', inspection.center_name)
        inspection.cost = request.POST.get('cost', inspection.cost) or 0
        if request.FILES.get('document'):
            inspection.document = request.FILES['document']
        inspection.save()
        messages.success(request, 'Visite technique modifiée avec succès !')
        return redirect('core:admin_inspections')

    context = {'inspection': inspection, 'cars': Car.objects.all()}
    return render(request, 'core/admin/inspection_form.html', context)


def admin_inspection_delete(request, inspection_id):
    """Supprimer une visite technique."""
    inspection = get_object_or_404(TechnicalInspection, id=inspection_id)
    if request.method == 'POST':
        inspection.delete()
        messages.success(request, 'Visite technique supprimée avec succès !')
    return redirect('core:admin_inspections')


# ============================================
#  ADMIN — ASSURANCES
# ============================================

def admin_insurances(request):
    """Liste des assurances par véhicule."""
    cars = Car.objects.all().prefetch_related('insurances')
    car_list = []
    today = datetime.date.today()
    for car in cars:
        latest = car.insurances.order_by('-expiry_date').first()
        is_valid = latest.expiry_date >= today if latest and latest.expiry_date else False
        car_list.append({
            'car': car,
            'latest_insurance': latest,
            'is_valid': is_valid
        })
    context = {'car_list': car_list}
    return render(request, 'core/admin/insurances_list.html', context)


def admin_insurance_add(request):
    """Ajouter une assurance."""
    if request.method == 'POST':
        insurance = Insurance(
            car_id=int(request.POST.get('car', 0)),
            company=request.POST.get('company', ''),
            insurance_type=request.POST.get('insurance_type', 'tous_risques'),
            policy_number=request.POST.get('policy_number', ''),
            start_date=request.POST.get('start_date'),
            expiry_date=request.POST.get('expiry_date'),
            annual_premium=request.POST.get('annual_premium', 0) or 0,
        )
        if request.FILES.get('document'):
            insurance.document = request.FILES['document']
        insurance.save()
        messages.success(request, 'Assurance ajoutée avec succès !')
        return redirect('core:admin_insurances')

    context = {
        'cars': Car.objects.all(),
        'insurance_types': Insurance.TYPE_CHOICES,
        'selected_car_id': request.GET.get('car'),
    }
    return render(request, 'core/admin/insurance_form.html', context)


def admin_insurance_edit(request, insurance_id):
    """Modifier une assurance."""
    insurance = get_object_or_404(Insurance, id=insurance_id)
    if request.method == 'POST':
        insurance.car_id = int(request.POST.get('car', insurance.car_id))
        insurance.company = request.POST.get('company', insurance.company)
        insurance.insurance_type = request.POST.get('insurance_type', insurance.insurance_type)
        insurance.policy_number = request.POST.get('policy_number', insurance.policy_number)
        insurance.start_date = request.POST.get('start_date', insurance.start_date)
        insurance.expiry_date = request.POST.get('expiry_date', insurance.expiry_date)
        insurance.annual_premium = request.POST.get('annual_premium', insurance.annual_premium) or 0
        if request.FILES.get('document'):
            insurance.document = request.FILES['document']
        insurance.save()
        messages.success(request, 'Assurance modifiée avec succès !')
        return redirect('core:admin_insurances')

    context = {
        'insurance': insurance,
        'cars': Car.objects.all(),
        'insurance_types': Insurance.TYPE_CHOICES,
    }
    return render(request, 'core/admin/insurance_form.html', context)


def admin_insurance_delete(request, insurance_id):
    """Supprimer une assurance."""
    get_object_or_404(Insurance, id=insurance_id).delete()
    messages.success(request, 'Assurance supprimée avec succès !')
    return redirect('core:admin_insurances')


# ============================================
#  ADMIN — AVIS CLIENTS
# ============================================

def admin_reviews(request):
    """Liste des avis clients."""
    all_reviews = Review.objects.all()
    context = {'all_reviews': all_reviews}
    return render(request, 'core/admin/reviews_list.html', context)


def admin_review_approve(request, review_id):
    """Approuver un avis."""
    review = get_object_or_404(Review, id=review_id)
    review.is_approved = not review.is_approved
    review.save()
    status = "approuvé" if review.is_approved else "masqué"
    messages.success(request, f'Avis de "{review.name}" {status}.')
    return redirect('core:admin_reviews')


def admin_review_delete(request, review_id):
    """Supprimer un avis."""
    review = get_object_or_404(Review, id=review_id)
    name = review.name
    review.delete()
    messages.success(request, f'Avis de "{name}" supprimé.')
    return redirect('core:admin_reviews')


# ============================================
#  ADMIN — MESSAGES
# ============================================

def admin_messages(request):
    """Liste des messages de contact."""
    all_messages = ContactMessage.objects.all()
    context = {'all_messages': all_messages}
    return render(request, 'core/admin/messages_list.html', context)


def admin_message_read(request, msg_id):
    """Marquer un message comme lu/non lu."""
    msg = get_object_or_404(ContactMessage, id=msg_id)
    msg.is_read = not msg.is_read
    msg.save()
    return redirect('core:admin_messages')


def admin_message_delete(request, msg_id):
    """Supprimer un message."""
    msg = get_object_or_404(ContactMessage, id=msg_id)
    msg.delete()
    messages.success(request, 'Message supprimé.')
    return redirect('core:admin_messages')
